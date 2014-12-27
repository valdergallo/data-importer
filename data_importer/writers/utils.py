# -*- encoding: utf-8 -*-
import zipfile
import re
from openpyxl import Workbook
from openpyxl import styles
from cStringIO import StringIO
from django.core.files import File
from django.utils.safestring import mark_safe
from django.utils.datastructures import SortedDict
import unicodedata

try:
    from django.utils import timezone
except ImportError:
    from datetime import datetime as timezone

# default yellow backgroung for first line
DEFAULT_HEADER_STYLE = styles.Style(fill=styles.PatternFill(patternType='solid', fgColor=styles.colors.YELLOW))


def slugify(value):
    """
    Converts to ASCII. Converts spaces to hyphens. Removes characters that
    aren't alphanumerics, underscores, or hyphens. Converts to lowercase.
    Also strips leading and trailing whitespace.
    """
    value = unicodedata.normalize('NFKD', unicode(value)).encode('ascii', 'ignore').decode('ascii')
    value = re.sub('[^\w\s-]', '', value).strip().lower()
    return mark_safe(re.sub('[-\s]+', '_', value))


class QuerysetToWorkbook(object):

    def __init__(self, queryset, columns, filename='report', header_style=DEFAULT_HEADER_STYLE):
        self.workbook = Workbook(guess_types=True)
        self.sheet = self.workbook.active
        self.queryset = queryset
        self.columns = columns
        self.style = header_style
        self.filename = filename

    def get_colums(self):
        return self._columns

    def set_columns(self, values):
        if isinstance(values, dict):
            self._columns = SortedDict(values)
        if isinstance(values, tuple) or isinstance(values, list):
            keys = [slugify(i) for i in values]
            self._columns = SortedDict(zip(keys, values))

    columns = property(get_colums, set_columns)

    def get_first(self, row=1):
        dimensions = str(self.sheet.dimensions)
        letters = re.findall('[A-Z]', dimensions)
        first = letters.pop(0)
        second = ''.join(letters)
        return '%s1:%s1' % (first, second)

    def set_header_style(self, row=1):
        try:
            row = self.sheet.row_dimensions[row]
            row.style = self.style
        except (KeyError, AttributeError):
            return

        for row in self.sheet.iter_rows(self.get_first(row)):
            for cell in row:
                cell.style = self.style

    def get_data(self, columns, data):
        line_dict = SortedDict()
        for k in columns.keys():
            line_dict[k] = getattr(data, k)
        return line_dict

    def queryset_to_workbook(self):
        if isinstance(self.columns, list) or isinstance(self.columns, tuple):
            self.sheet.append(self.columns.values())
        for line in self.queryset:
            if hasattr(line, '__dict__'):
                data = self.get_data(self.columns, line)
                self.sheet.append(data.values())
            elif isinstance(line, list) or isinstance(line, tuple):
                self.sheet.append(line)
        return self.workbook

    def get_content(self):
        in_memory = StringIO()
        wb = self.queryset_to_workbook()
        self.set_header_style()
        wb.save(in_memory)
        return in_memory

    def get_compressed_file(self):
        in_memory_zip = StringIO()
        zf = zipfile.ZipFile(in_memory_zip, "w", zipfile.ZIP_DEFLATED)
        zf.writestr(self.get_filename(), self.get_content().getvalue())
        zf.close()
        in_memory_zip.flush()
        return in_memory_zip

    def get_filename(self, extension='xlsx', username=None):
        if username:
            return self.filename + '_' + username + '_' + timezone.now().strftime('%Y%m%d') + '.' + extension
        return self.filename + '_' + timezone.now().strftime('%Y%m%d') + '.' + extension

    def compress_django_file(self, compress=True, filename=None):
        return File(self.get_compressed_file())

    def save(self, compress=True, filename=None):
        if compress:
            if not filename:
                filename = self.get_filename(extension='zip')
            with open(filename, 'wb') as fl:
                fl.write(self.get_compressed_file().getvalue())
                return fl
        else:
            if not filename:
                filename = self.get_filename()
            with open(filename, 'wb') as fl:
                fl.write(self.get_content().getvalue())
                return fl

    def response(self, compress=True, filename=None):
        from django.http import HttpResponse
        from django.core.servers.basehttp import FileWrapper

        if compress:
            response = HttpResponse(FileWrapper(self.get_compressed_file()))
            _filename = self.get_filename(extension='zip')
        else:
            response = HttpResponse(FileWrapper(self.get_content()))
            _filename = self.get_filename()

        if not filename:
            filename = _filename

        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        response['content_type'] = 'application/force-download'
        return response
