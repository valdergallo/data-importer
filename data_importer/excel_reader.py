# coding: utf-8
import datetime
import xlrd
import openpyxl
from openpyxl import load_workbook
from base import BaseImporter

#from xlrd/biffh.py
(
    XL_CELL_EMPTY,
    XL_CELL_TEXT,
    XL_CELL_NUMBER,
    XL_CELL_DATE,
    XL_CELL_BOOLEAN,
    XL_CELL_ERROR,
    XL_CELL_BLANK,  # for use in debugging, gathering stats, etc
) = range(7)


class XLSReader(BaseImporter):

    def set_reader(self, on_demand=True):
        self.workbook = xlrd.open_workbook(self.source, on_demand=on_demand)
        if self.Meta.sheet_name:
            self.worksheet = self.workbook.sheet_by_name(self.Meta.sheet_name)
        else:
            self.worksheet = self.workbook.sheet_by_index(0)

        self._reader = self.get_items()

    def convert_value(self, item):
        """
        Handle different value types for XLS. Item is a cell object.
        """
        # Thx to Augusto C Men to point fast solution for XLS/XLSX dates
        if item.ctype == 3: #XL_CELL_DATE:
            return datetime.datetime(*xlrd.xldate_as_tuple(item.value, self.workbook.datemode))

        if item.ctype == 2: #XL_CELL_NUMBER:
            if item.value % 1 == 0:  # integers
                return int(item.value)
            else:
                return item.value

        return item.value

    def get_items(self):
        start_line = 0
        if self.Meta.ignore_first_line:
            start_line = 1

        for i in range(start_line, self.worksheet.nrows):
            values = [self.convert_value(cell) for cell in self.worksheet.rows[i]]
            if not any(values):
                continue  # empty lines are ignored
            yield values


class XLSXReader(XLSReader):

    def set_reader(self, use_iterators=True):
        self.workbook = load_workbook(self.source, use_iterators=use_iterators)
        if self.Meta.sheet_name:
            self.worksheet = self.workbook.get_sheet_by_name(self.Meta.sheet_name)
        else:
            self.worksheet = self.workbook.worksheets[0]

        self._reader = self.get_items()

    def get_value(self, item):
        """
        Handle different value types for XLSX. Item is a cell object.
        """
        if item.is_date:
            return item.internal_value # return datetime

        if item.internal_value % 1 == 0:  # return integers
            return int(item.internal_value)

        return item.internal_value # return string

    def get_items(self):
        for line, row in enumerate(self.worksheet.iter_rows()):
            if self.Meta.ignore_first_line and line == 0:
                pass
            else:
                values = [self.get_value(cell) for cell in row]
                if not any(values):
                    continue  # empty lines are ignored
                yield self.get_item(values)
