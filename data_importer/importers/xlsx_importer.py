#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from .base import BaseImporter


class XLSXImporter(BaseImporter):

    def set_reader(self, use_iterators=True, data_only=True):
        "Read XLSX files"
        self.workbook = load_workbook(self.source,
                                      use_iterators=use_iterators,
                                      data_only=data_only)
        if self.Meta.sheet_name:
            self.worksheet = self.workbook.get_sheet_by_name(self.Meta.sheet_name)
        else:
            self.worksheet = self.workbook.worksheets[0]

        self._reader = self.get_items()

    def convert_value(self, item):
        """
        Handle different value types for XLSX. Item is a cell object.
        """
        return item.value

    def get_items(self):
        """
        Get values from cells
        :return: generator
        """
        for line, row in enumerate(self.worksheet.iter_rows()):
            values = [self.convert_value(cell) for cell in row]
            if not any(values):
                continue  # empty lines are ignored
            yield values
