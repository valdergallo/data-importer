#!/usr/bin/python
# encoding: utf-8
import datetime
import openpyxl
from openpyxl import load_workbook
from .base import BaseImporter

#from xlrd/biffh.py
(
    XL_CELL_EMPTY,
    XL_CELL_TEXT,
    XL_CELL_NUMBER,
    XL_CELL_DATE,
    XL_CELL_BOOLEAN,
    XL_CELL_ERROR,
    XL_CELL_BLANK,  # for use in debugging, gathering stats, etc
) = range(7)



class XLSXReader(XLSReader):

    def set_reader(self, use_iterators=True):
        self.workbook = load_workbook(self.source, use_iterators=use_iterators)
        if self.Meta.sheet_name:
            self.worksheet = self.workbook.get_sheet_by_name(self.Meta.sheet_name)
        else:
            self.worksheet = self.workbook.worksheets[0]

        self._reader = self.get_items()

    def get_value(self, item):
        """
        Handle different value types for XLSX. Item is a cell object.
        """
        if item.is_date:
            return item.internal_value # return datetime

        if item.internal_value % 1 == 0:  # return integers
            return int(item.internal_value)

        return item.internal_value # return string

    def get_items(self):
        for line, row in enumerate(self.worksheet.iter_rows()):
            if self.Meta.ignore_first_line and line == 0:
                pass
            else:
                values = [self.get_value(cell) for cell in row]
                if not any(values):
                    continue  # empty lines are ignored
                yield self.get_item(values)
