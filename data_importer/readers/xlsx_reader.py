# encoding: utf-8
from __future__ import unicode_literals
from openpyxl import load_workbook


class XLSXReader(object):

    def __init__(self, instance, data_only=True, sheet_index=None):
        self.workbook = load_workbook(instance.source,
                                      data_only=data_only)
        if instance.Meta.sheet_name:
            self.worksheet = self.workbook.get_sheet_by_name(instance.Meta.sheet_name)
        else:
            if sheet_index is None:
                sheet_index = instance.Meta.sheet_index or 0
            self.worksheet = self.workbook.worksheets[sheet_index]

    def read(self):
        for line, row in enumerate(self.worksheet.iter_rows()):
            values = [cell.value for cell in row]
            yield values
